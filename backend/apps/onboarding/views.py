import logging

from django.utils import timezone
from django.db.models import Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from .models import (
    Onboarding, OnboardingToken, VendorReferenceMaster, PaymentTermMaster,
    PurchaseOrganizationMaster, CompanyCodeMaster, TDSCodeMaster,
    SearchTermMaster, OnboardingApprovalHistory,
    SalesOrganizationMaster, DistributionChannelMaster, DivisionMaster,
    TransportationZoneMaster, CustomerCompanyCodeMaster,
    VENDOR_REFERENCE_RANGES, get_vendor_reference_range_for_code,
)
from .serializers import (
    OnboardingListSerializer, OnboardingDetailSerializer,
    CreateOnboardingSerializer, ApproveRejectSerializer,
    OnboardingTokenSerializer, VendorReferenceMasterSerializer,
    VendorReferenceLookupSerializer, gst_state_code_for_state,
    pan_name_is_editable,
)
from apps.accounts.models import User
from apps.notifications.email_service import send_boss_approval_request, send_onboarding_invite
from config import settings


PENDING_STATUS_FILTER = ['PENDING', 'PENDING_BOSS_APPROVAL', 'UNDER_REVIEW']
logger = logging.getLogger(__name__)


def _notify_boss_of_approval_request(onboarding, boss, employee):
    """Email notification failures must not undo a successfully submitted request."""
    try:
        send_boss_approval_request(boss.email, onboarding, employee)
    except Exception:
        logger.exception(
            'Could not send approval notification for onboarding %s to %s',
            onboarding.id,
            boss.email,
        )


def _filter_by_created_date(qs, query_params):
    """Filter a queryset inclusively by ISO-8601 created-date query parameters."""
    start_date_raw = query_params.get('start_date')
    end_date_raw = query_params.get('end_date')
    start_date = parse_date(start_date_raw) if start_date_raw else None
    end_date = parse_date(end_date_raw) if end_date_raw else None

    errors = {}
    if start_date_raw and not start_date:
        errors['start_date'] = 'Use the YYYY-MM-DD format.'
    if end_date_raw and not end_date:
        errors['end_date'] = 'Use the YYYY-MM-DD format.'
    if errors:
        raise ValueError(errors)
    if start_date and end_date and start_date > end_date:
        raise ValueError({'end_date': 'End date must be on or after start date.'})

    if start_date:
        qs = qs.filter(created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(created_at__date__lte=end_date)
    return qs


class IsAdmin(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and request.user.role in ['ADMIN', 'BOSS', 'EMPLOYEE']


class IsFullAdmin(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and _is_system_admin(request.user)


class IsAdminOrBoss(IsAuthenticated):
    def has_permission(self, request, view):
        return super().has_permission(request, view) and (
            request.user.is_superuser or request.user.role in ['ADMIN', 'BOSS']
        )


def _is_system_admin(user):
    return bool(user.is_superuser or user.role == 'ADMIN')


# ── Scope helpers ────────────────────────────────────────────────
def _scoped_qs(user):
    """Onboardings visible to this user based on their role."""
    qs = Onboarding.objects.select_related('created_by', 'approved_by', 'assigned_user').prefetch_related('created_by__bosses').all()
    if _is_system_admin(user):
        return qs
    if user.role == 'BOSS':
        return qs.filter(Q(assigned_user=user) | Q(created_by=user)).distinct()
    if user.role == 'EMPLOYEE':
        return qs.filter(created_by=user)
    return qs.filter(assigned_user=user)


def _can_access(user, onboarding):
    """Whether user may view/edit a specific onboarding."""
    if _is_system_admin(user):
        return True
    if user.role == 'BOSS':
        return (
            str(onboarding.created_by_id) == str(user.id)
            or str(onboarding.assigned_user_id) == str(user.id)
        )
    if user.role == 'EMPLOYEE':
        return str(onboarding.created_by_id) == str(user.id)
    return str(onboarding.assigned_user_id) == str(user.id)


def _can_edit_onboarding(user, onboarding):
    if _is_system_admin(user):
        return True
    if user.role in ['BOSS', 'EMPLOYEE']:
        return _can_access(user, onboarding)
    return False


def _bosses_for_user(user):
    if user.role == 'EMPLOYEE':
        return list(user.bosses.all())
    return []


def _selected_approval_boss(user, data):
    if user.role != 'EMPLOYEE':
        return None, None

    boss_id = data.get('approval_boss') or data.get('assigned_boss') or data.get('assigned_user')
    bosses = user.bosses.all()
    if not bosses.exists():
        return None, {'approval_boss': ['Employee is not assigned to any boss.']}
    if not boss_id:
        if bosses.count() == 1:
            return bosses.first(), None
        return None, {'approval_boss': ['Select the boss to send this request for approval.']}

    selected_boss = bosses.filter(id=boss_id).first()
    if not selected_boss:
        return None, {'approval_boss': ['Selected boss is not assigned to this employee.']}
    return selected_boss, None


def _validate_required_manual_onboarding_fields(data):
    errors = {}
    is_customer = str(data.get('onboarding_type', '') or '').strip().upper() == 'CUSTOMER'
    required_text_fields = {
        'company_name': 'Company name is required.',
        'city': 'City is required.',
        'state': 'State is required.',
        'street1': 'Street address is required.',
        'pan_number': 'PAN number is required.',
    }
    if not is_customer:
        required_text_fields.update({
            'account_holder_name': 'Account holder name is required.',
            'bank_name': 'Bank name is required.',
            'branch_name': 'Branch name is required.',
            'account_number': 'Account number is required.',
            'ifsc_code': 'IFSC code is required.',
        })
    for field, message in required_text_fields.items():
        if not str(data.get(field, '')).strip():
            errors[field] = [message]

    if pan_name_is_editable(data.get('pan_number')) and not str(data.get('pan_name', '')).strip():
        errors['pan_name'] = ['PAN name is required.']

    if not data.get('emails'):
        errors['emails'] = ['At least one email is required.']
    if not data.get('phones'):
        errors['phones'] = ['At least one phone number is required.']

    pincode = str(data.get('pincode', '')).strip()
    if len(pincode) != 6:
        errors['pincode'] = ['6-digit PIN code is required.']

    if data.get('gst_applicable') is None:
        errors['gst_applicable'] = ['Please select GST status.']
    elif data.get('gst_applicable') and not str(data.get('gst_number', '')).strip():
        errors['gst_number'] = ['GST number is required.']

    if not is_customer:
        if data.get('msme_applicable') is None:
            errors['msme_applicable'] = ['Please select MSME status.']
        elif data.get('msme_applicable'):
            if not str(data.get('msme_category', '')).strip():
                errors['msme_category'] = ['MSME category is required.']
            if not str(data.get('udyam_number', '')).strip():
                errors['udyam_number'] = ['Udyam registration number is required.']

    return errors


def _validate_required_onboarding_documents(onboarding):
    existing_types = set(onboarding.documents.values_list('document_type', flat=True))
    is_customer = str(getattr(onboarding, 'onboarding_type', '') or '').strip().upper() == 'CUSTOMER'
    errors = {}
    if 'PAN' not in existing_types:
        errors['pan_doc'] = ['PAN Card document is required.']
    if onboarding.gst_applicable and 'GST' not in existing_types:
        errors['gst_doc'] = ['GST Certificate is required.']
    if not is_customer:
        if 'CHEQUE' not in existing_types:
            errors['cheque_doc'] = ['Cancelled cheque is required.']
        if onboarding.msme_applicable and 'MSME' not in existing_types:
            errors['msme_doc'] = ['MSME Certificate is required.']
    return errors


def _entity_type_for_onboarding(onboarding):
    onboarding_code = str(getattr(onboarding, 'onboarding_code', '') or '').strip().upper()
    if onboarding_code.startswith('V'):
        return 'Vendor'
    if onboarding_code.startswith('C'):
        return 'Customer'

    onboarding_type = str(getattr(onboarding, 'onboarding_type', '') or '').strip().upper()
    if onboarding_type == 'VENDOR':
        return 'Vendor'
    if onboarding_type == 'CUSTOMER':
        return 'Customer'

    return 'Business Partner'


PAN_APPROVAL_PENDING = 'pending'
PAN_APPROVAL_VALID_OPERATIVE = 'valid_operative'
PAN_APPROVAL_VALID_INOPERATIVE = 'valid_inoperative'
PAN_APPROVAL_FAILED = 'failed'
GST_APPROVAL_PENDING = 'pending'
GST_APPROVAL_VALID = 'valid'
GST_APPROVAL_FAILED = 'failed'
GST_APPROVAL_NOT_APPLICABLE = 'not_applicable'


def _has_any(value, terms):
    normalized = str(value or '').lower()
    return any(term in normalized for term in terms)


def _classify_pan_approval_status(onboarding):
    verification_status = onboarding.pan_verification_status
    if not onboarding.pan_number or not verification_status:
        return PAN_APPROVAL_PENDING
    if _has_any(verification_status, ['invalid', 'failed', 'failure', 'error', 'no records', 'not found']):
        return PAN_APPROVAL_FAILED
    if _has_any(verification_status, ['inoperative', 'not operative']):
        return PAN_APPROVAL_VALID_INOPERATIVE
    if onboarding.pan_verified or _has_any(verification_status, ['valid']):
        return PAN_APPROVAL_VALID_OPERATIVE
    return PAN_APPROVAL_FAILED


def _classify_gst_approval_status(onboarding):
    if not onboarding.gst_applicable:
        return GST_APPROVAL_NOT_APPLICABLE
    verification_status = onboarding.gst_verification_status
    if not onboarding.gst_number or not verification_status:
        return GST_APPROVAL_PENDING
    if (
        _has_any(verification_status, ['invalid', 'failed', 'failure', 'error', 'no records', 'not found', 'cancelled', 'suspended'])
        or _has_any(verification_status, ['found but status'])
    ):
        return GST_APPROVAL_FAILED
    if onboarding.gst_verified or _has_any(verification_status, ['valid']):
        return GST_APPROVAL_VALID
    return GST_APPROVAL_FAILED


def _validate_approval_verifications(onboarding):
    pan_status = _classify_pan_approval_status(onboarding)
    gst_status = _classify_gst_approval_status(onboarding)
    gst_passed = gst_status in (GST_APPROVAL_VALID, GST_APPROVAL_NOT_APPLICABLE)

    if pan_status == PAN_APPROVAL_VALID_OPERATIVE and gst_passed:
        return []

    if pan_status == PAN_APPROVAL_PENDING and gst_status == GST_APPROVAL_PENDING:
        return ['Cannot approve vendor. Both PAN and GST verifications are pending. Please complete both verifications before approval.']
    if pan_status == PAN_APPROVAL_FAILED and gst_status == GST_APPROVAL_FAILED:
        return ['Cannot approve vendor. Both PAN and GST verification have failed. Please correct the details and verify again.']
    if pan_status == PAN_APPROVAL_FAILED and gst_status == GST_APPROVAL_PENDING:
        return ['Cannot approve vendor. PAN verification has failed and GST verification is still pending. Please correct the PAN details and complete GST verification.']
    if pan_status == PAN_APPROVAL_PENDING and gst_status == GST_APPROVAL_FAILED:
        return ['Cannot approve vendor. GST verification has failed and PAN verification is still pending. Please complete PAN verification and correct the GST details.']

    messages = []
    if pan_status == PAN_APPROVAL_PENDING:
        messages.append('Cannot approve vendor. PAN verification is pending. Please verify the PAN first.')
    elif pan_status == PAN_APPROVAL_FAILED:
        messages.append('Cannot approve vendor. PAN verification failed. Please correct the PAN details and verify again.')
    elif pan_status == PAN_APPROVAL_VALID_INOPERATIVE:
        messages.append('Cannot approve vendor. The PAN is valid but currently inoperative. Please provide an operative PAN before approval.')

    if gst_status == GST_APPROVAL_PENDING:
        messages.append('Cannot approve vendor. GST verification is pending. Please verify the GST first.')
    elif gst_status == GST_APPROVAL_FAILED:
        messages.append('Cannot approve vendor. GST verification failed. Please correct the GST details and verify again.')

    return messages


# ── Admin: List onboardings ──────────────────────────────────────
class OnboardingListView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        qs = _scoped_qs(request.user)
        try:
            qs = _filter_by_created_date(qs, request.query_params)
        except ValueError as error:
            return Response(error.args[0], status=status.HTTP_400_BAD_REQUEST)

        # Filter by type
        onboarding_type = request.query_params.get('type')
        if onboarding_type:
            qs = qs.filter(onboarding_type=onboarding_type.upper())

        # Filter by status
        status_filter = request.query_params.get('status')
        if status_filter:
            normalized_status = status_filter.upper()
            if normalized_status == 'PENDING_GROUP':
                qs = qs.filter(status__in=PENDING_STATUS_FILTER)
            else:
                qs = qs.filter(status=normalized_status)

        # Search
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(company_name__icontains=search) |
                Q(onboarding_code__icontains=search) |
                Q(pan_number__icontains=search) |
                Q(contact_person__icontains=search)
            )

        serializer = OnboardingListSerializer(qs, many=True)
        return Response(serializer.data)


def _fmt_dt(value):
    return timezone.localtime(value).strftime('%d-%m-%Y %H:%M') if value else ''


def _fmt_date(value):
    return value.isoformat() if value else ''


def _yes_no(value):
    return 'Yes' if value else 'No'


SAP_REGION_CODES = {
    'Andaman & Nicobar Islands': 'AN',
    'Andaman And Nicobar': 'AN',
    'Andhra Pradesh': 'AP',
    'Arunachal Pradesh': 'ARP',
    'Assam': 'AS',
    'Bihar': 'BH',
    'Chandigarh': 'CD',
    'Chhattisgarh': 'CH',
    'Chattisgarh': 'CH',
    'Dadra And Nagar Haveli': 'DH',
    'Dadra & Nagar Haveli': 'DH',
    'Daman & Diu': 'DD',
    'Delhi': 'DEL',
    'Goa': 'GDD',
    'Gujarat': 'GJ',
    'Haryana': 'HR',
    'Himachal Pradesh': 'HP',
    'Jammu And Kashmir': 'JK',
    'Jammu & Kashmir': 'JK',
    'Jharkhand': 'JH',
    'Karnataka': 'KAR',
    'Kerala': 'KER',
    'Lakshadweep': 'LAK',
    "Ladakh": "LA",
    'Madhya Pradesh': 'MP',
    'Maharashtra': 'MAH',
    'Meghalaya': 'MG',
    'Mizoram': 'MZ',
    'Nagaland': 'NG',
    'Odisha': 'OR',
    'Orissa': 'OR',
    'Puducherry': 'PY',
    'Pondicherry': 'PY',
    'Punjab': 'PB',
    'Rajasthan': 'RJ',
    'Sikkim': 'SK',
    'Tamil Nadu': 'TN',
    'Telangana': 'TG',
    'Tripura': 'TP',
    'Uttarakhand': 'UT',
    'Uttaranchal': 'UT',
    'Uttar Pradesh': 'UP',
    'West Bengal': 'WB',
    "Export": 'EXP'
}


def _first_list_value(value):
    if isinstance(value, list):
        return str(value[0] or '').strip() if value else ''
    return str(value or '').split(',')[0].strip()


def _name_part(value, index, size=35):
    normalized = ' '.join(str(value or '').split())
    start = index * size
    return normalized[start:start + size]


def _sap_region_code(state):
    normalized = str(state or '').strip()
    if not normalized:
        return ''
    return SAP_REGION_CODES.get(normalized, normalized.upper())


def _first_tds_code(value):
    parts = _first_list_value(value).split()
    return parts[0].strip() if parts else ''


def _bank_account_main(value):
    return str(value or '').strip()[:18]


def _bank_account_extra(value):
    account_number = str(value or '').strip()
    return account_number[18:] if len(account_number) > 18 else ''


def _vendor_master(onboarding):
    return getattr(onboarding, '_vendor_reference_master', None)


def _vendor_group_code(onboarding):
    master = _vendor_master(onboarding)
    return master.group_code if master else ''


class OnboardingExportView(APIView):
    permission_classes = [IsAdmin]

    COLUMNS = [
        ('Partner', lambda o: ''),
        ('BP_ROLE', lambda o: 'GEN'),
        ('CREATION_GROUP', lambda o: 'FLVN010X'),
        ('PARTN_GRP', _vendor_group_code),
        ('SEARCHTERM1', lambda o: o.search_term),
        ('SEARCHTERM2', lambda o: o.pan_number),
        ('PARTNERTYPE', lambda o: ''),
        ('AUTHORIZATIONGROUP', lambda o: ''),
        ('PARTNERLANGUAGE', lambda o: ''),
        ('TITLE_KEY', lambda o: ''),
        ('NAME1', lambda o: _name_part(o.company_name, 0)),
        ('NAME2', lambda o: _name_part(o.company_name, 1)),
        ('NAME3', lambda o: _name_part(o.company_name, 2)),
        ('NAME4', lambda o: _name_part(o.company_name, 3)),
        ('INDUSTRYSECTOR', lambda o: ''),
        ('CITY', lambda o: o.city),
        ('STREET', lambda o: o.street1),
        ('STR_SUPPL1 (Street 2)', lambda o: o.street2),
        ('STR_SUPPL2(Street3)', lambda o: o.street3),
        ('STR_SUPPL3(Street 4)', lambda o: o.street4),
        ('LOCATION', lambda o: o.district),
        ('POSTL_COD1(Postal Code)', lambda o: o.pincode),
        ('COUNTRY', lambda o: 'IN' if str(o.country or '').strip().lower() in ['', 'india'] else o.country),
        ('REGION', lambda o: _sap_region_code(o.state)),
        ('TRANSPZONE', lambda o: ''),
        ('TAXTYPE', lambda o: ''),
        ('TAXNUMBER (GST)', lambda o: o.gst_number),
        ('ZUAWA(Sort Key)', lambda o: ''),
        ('AKONT(Reconciliation acct)', lambda o: o.gl_account_number),
        ('BEGRU(Auth. Grp.)', lambda o: ''),
        ('ZWELS(Payment Method)', lambda o: 'ACHORSXYZ'),
        ('ZTERM (payment term)', lambda o: o.payment_terms),
        ('REPRF(Check Double Invoice)', lambda o: 'X'),
        ('TOGRU(Tolerance grp)', lambda o: ''),
        ('HBKID (house Bank)', lambda o: ''),
        ('QSZNR(Exemption Number)', lambda o: ''),
        ('QSZDT(Valid Until)', lambda o: ''),
        ('QSSKZ(WTax Code)', lambda o: ''),
        ('MINDK(Minority Indic.)', lambda o: o.msme_status),
        ('UZAWE(Pmt Meth. Supplement)', lambda o: ''),
        ('TLFNS(Acct.clerks tel.no.)', lambda o: ''),
        ('EKORG(Purchasing Org.)', lambda o: _first_list_value(o.purchase_orgs_to_open or o.reference_purchase_orgs)),
        ('WAERS(Currency)', lambda o: 'INR'),
        ('VZTERM(Payment term key)', lambda o: ''),
        ('INCO1(Incoterms)', lambda o: ''),
        ('INCO2_L(Incoterms1)', lambda o: ''),
        ('INCO3_L(Incoterms2)', lambda o: ''),
        ('WEBRE(GR-Based Inv. Verif.)', lambda o: 'X'),
        ('LEBRE(Service-Based Invoice Verification)', lambda o: 'X'),
        ('EKGRP(Purchase Grp)', lambda o: ''),
        ('KALSK(Schema Grp Supp)', lambda o: ''),
        ('BUKRS(Comp code)', lambda o: o.company_code_to_open),
        ('PANNO', lambda o: o.pan_number),
        ('TELEPHONE', lambda o: _first_list_value(o.phones)),
        ('EXTENSION', lambda o: ''),
        ('TEL_NO(Telephone2)', lambda o: _first_list_value(o.phones)),
        ('EXTENSION2', lambda o: ''),
        ('E_MAIL', lambda o: _first_list_value(o.emails)),
        ('BANK_CTRY(Bank COuntry)', lambda o: 'IN'),
        ('BANK_KEY (IFSC Code)', lambda o: o.ifsc_code),
        ('BANK_ACCT(Bank Account)', lambda o: _bank_account_main(o.account_number)),
        ('Reference Details (A/C No-extra digits if longer than 18 chars.)', lambda o: _bank_account_extra(o.account_number)),
        ('ACCOUNTHOLDER', lambda o: o.account_holder_name),
        ('BANKACCOUNTNAME', lambda o: ''),
        ('BP_BANK_GUID', lambda o: ''),
        ('WITHT(Withholding Tax Type)', lambda o: _first_tds_code(o.tds_codes)),
        ('WT_SUBJCT(Subject to w/tax)', lambda o: 'X' if _first_tds_code(o.tds_codes) else ''),
        ('QSREC(Recipient Type)', lambda o: 'OT' if _first_tds_code(o.tds_codes) else ''),
        ('WT_WTSTCD(W/tax identification no.)', lambda o: ''),
        ('WT_WITHCD(W/Tax Code)', lambda o: _first_tds_code(o.tds_codes)),
        ('WT_EXNR(Exemption Number)', lambda o: ''),
        ('WT_EXRT(Exemption Rate)', lambda o: ''),
        ('WT_EXDF(Exemption Start Date)', lambda o: ''),
        ('WT_EXDT(Exemption End Date)', lambda o: ''),
        ('WT_WTEXRS(Exempt. Reason)', lambda o: ''),
        ('Name of Representative', lambda o: ''),
        ('Type of Industry', lambda o: ''),
        ('Planning Group', lambda o: 'O2'),
        ('ServAgntProcGrp(DLGRP)', lambda o: ''),
        ('Stat.grp, agent', lambda o: ''),
        ('Shipping Conditions', lambda o: ''),
        ('Part_categr(1,2 or3)', lambda o: 2),
        ('Certification date', lambda o: ''),
        ('Vendor_class', lambda o: 0),
    ]

    def build_workbook_response(self, queryset, filename):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        vendor_masters = {
            master.vendor_reference_range: master
            for master in VendorReferenceMaster.objects.all()
        }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Mass Vendor Creation Template'
        worksheet.append([header for header, _ in self.COLUMNS])

        for onboarding in queryset:
            onboarding._vendor_reference_master = vendor_masters.get(onboarding.vendor_reference_range)
            worksheet.append([getter(onboarding) for _, getter in self.COLUMNS])

        for index in range(1, len(self.COLUMNS) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = 22
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        worksheet.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response

    def get(self, request):
        try:
            queryset = _filter_by_created_date(_scoped_qs(request.user), request.query_params)
        except ValueError as error:
            return Response(error.args[0], status=status.HTTP_400_BAD_REQUEST)

        queryset = queryset.filter(status='APPROVED')
        return self.build_workbook_response(queryset, 'onboarding_export.xlsx')


class OnboardingSingleExportView(OnboardingExportView):
    permission_classes = [IsAdmin]

    def get(self, request, pk):
        queryset = _scoped_qs(request.user).filter(pk=pk, status='APPROVED')
        onboarding = queryset.first()
        if not onboarding:
            return Response(
                {'detail': 'Only approved records can be exported.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        filename = f'{onboarding.onboarding_code or "onboarding"}_export.xlsx'
        return self.build_workbook_response(queryset, filename)


class PanDataExportView(APIView):
    """Export PAN-specific data for records that have a PAN number."""
    permission_classes = [IsAdmin]

    def get(self, request):
        try:
            queryset = _filter_by_created_date(_scoped_qs(request.user), request.query_params)
        except ValueError as error:
            return Response(error.args[0], status=status.HTTP_400_BAD_REQUEST)

        from openpyxl import Workbook
        from openpyxl.styles import Font

        queryset = queryset.exclude(pan_number__isnull=True).exclude(pan_number='')
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'PAN Data'
        worksheet.append(['PAN Number'])

        for onboarding in queryset:
            worksheet.append([onboarding.pan_number])

        worksheet.column_dimensions['A'].width = 16
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="pan_data_export.xlsx"'
        workbook.save(response)
        return response


# ── Admin: Manual onboarding (no email required) ────────────────
class ManualOnboardingView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        onboarding_type = request.data.get('onboarding_type', '').upper()
        if onboarding_type not in ['VENDOR', 'CUSTOMER']:
            return Response(
                {'onboarding_type': ['Must be VENDOR or CUSTOMER.']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        data = {**request.data, 'onboarding_type': onboarding_type}
        required_errors = _validate_required_manual_onboarding_fields(data)
        if required_errors:
            return Response(required_errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = OnboardingDetailSerializer(
            data=data,
            partial=True,
            context={'require_complete': True},
        )
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Employee-created records must first complete their verifications and
        # required documents. They are sent to the boss only through the
        # guarded send-to-boss endpoint once every condition is green.
        next_status = 'DRAFT' if request.user.role == 'EMPLOYEE' else 'PENDING'
        onboarding = serializer.save(
            created_by=request.user,
            assigned_user=None,
            status=next_status,
        )
        OnboardingApprovalHistory.objects.create(
            onboarding=onboarding,
            action='SUBMITTED',
            actor=request.user,
            comments=(
                'Saved as draft. Complete all verifications and documents before sending to the boss.'
                if request.user.role == 'EMPLOYEE' else 'Submitted for approval.'
            ),
        )
        return Response(OnboardingDetailSerializer(onboarding).data, status=status.HTTP_201_CREATED)


# ── Admin: Bulk onboarding import from Excel ─────────────────────
BULK_IMPORT_COLUMNS = [
    ('company_name', 'Company Name'),
    ('contact_person', 'Contact Person'),
    ('emails', 'Email(s)'),
    ('phones', 'Phone(s)'),
    ('district', 'District'),
    ('city', 'City'),
    ('state', 'State'),
    ('pincode', 'Pincode'),
    ('country', 'Country'),
    ('street1', 'Street 1'),
    ('street2', 'Street 2'),
    ('street3', 'Street 3'),
    ('street4', 'Street 4'),
    ('pan_number', 'PAN Number'),
    ('gst_applicable', 'GST Applicable (Yes/No)'),
    ('gst_number', 'GST Number'),
    ('account_holder_name', 'Account Holder Name'),
    ('bank_name', 'Bank Name'),
    ('branch_name', 'Branch Name'),
    ('account_number', 'Account Number'),
    ('ifsc_code', 'IFSC Code'),
    ('msme_applicable', 'MSME Applicable (Yes/No)'),
    ('msme_category', 'MSME Category'),
    ('udyam_number', 'Udyam Number'),
]


class BulkOnboardingTemplateView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Bulk Onboarding Template'
        worksheet.append([header for _, header in BULK_IMPORT_COLUMNS])
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        for index in range(1, len(BULK_IMPORT_COLUMNS) + 1):
            worksheet.column_dimensions[worksheet.cell(row=1, column=index).column_letter].width = 24
        worksheet.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bulk_onboarding_template.xlsx"'
        workbook.save(response)
        return response


def _bulk_import_cell(value):
    if value is None:
        return ''
    return str(value).strip()


def _bulk_import_bool(value):
    normalized = _bulk_import_cell(value).lower()
    if normalized in ('yes', 'y', 'true', '1'):
        return True
    if normalized in ('no', 'n', 'false', '0', ''):
        return False
    return None


def _bulk_import_list(value):
    return [part.strip() for part in _bulk_import_cell(value).split(',') if part.strip()]


def _row_to_onboarding_data(row, headers, onboarding_type):
    padded_row = list(row) + [None] * (len(headers) - len(row))
    raw = dict(zip(headers, padded_row))
    data = {'onboarding_type': onboarding_type}
    for field, header in BULK_IMPORT_COLUMNS:
        data[field] = raw.get(header)

    data['emails'] = _bulk_import_list(data.get('emails'))
    data['phones'] = _bulk_import_list(data.get('phones'))
    data['country'] = _bulk_import_cell(data.get('country')) or 'India'
    data['pan_number'] = _bulk_import_cell(data.get('pan_number')).upper()
    data['gst_number'] = _bulk_import_cell(data.get('gst_number')).upper()
    data['ifsc_code'] = _bulk_import_cell(data.get('ifsc_code')).upper()
    data['pincode'] = _bulk_import_cell(data.get('pincode'))

    gst_applicable = _bulk_import_bool(data.get('gst_applicable'))
    data['gst_applicable'] = bool(gst_applicable)
    if not gst_applicable:
        data['gst_number'] = ''

    msme_applicable = _bulk_import_bool(data.get('msme_applicable'))
    data['msme_applicable'] = bool(msme_applicable)
    msme_category = _bulk_import_cell(data.get('msme_category')).upper()
    if msme_applicable:
        data['msme_category'] = msme_category
        data['msme_status'] = msme_category or 'MNA'
        data['udyam_number'] = _bulk_import_cell(data.get('udyam_number'))
    else:
        data['msme_category'] = ''
        data['msme_status'] = 'MNA'
        data['udyam_number'] = ''

    for field in ['company_name', 'contact_person', 'district', 'city', 'state',
                  'street1', 'street2', 'street3', 'street4',
                  'account_holder_name', 'bank_name', 'branch_name', 'account_number']:
        data[field] = _bulk_import_cell(data.get(field))

    return data


class BulkOnboardingImportView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        onboarding_type = request.data.get('onboarding_type', '').upper()
        if onboarding_type not in ['VENDOR', 'CUSTOMER']:
            return Response(
                {'onboarding_type': ['Must be VENDOR or CUSTOMER.']},
                status=status.HTTP_400_BAD_REQUEST,
            )

        upload = request.FILES.get('file')
        if not upload:
            return Response({'file': ['An Excel file is required.']}, status=status.HTTP_400_BAD_REQUEST)

        from openpyxl import load_workbook
        try:
            workbook = load_workbook(upload, data_only=True)
        except Exception:
            return Response({'file': ['Could not read the Excel file. Upload a valid .xlsx file.']}, status=status.HTTP_400_BAD_REQUEST)

        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return Response({'file': ['The file is empty.']}, status=status.HTTP_400_BAD_REQUEST)

        headers = [_bulk_import_cell(cell) for cell in rows[0]]
        data_rows = rows[1:]
        if not data_rows:
            return Response({'file': ['No data rows found below the header.']}, status=status.HTTP_400_BAD_REQUEST)

        created = []
        row_errors = []

        for index, row in enumerate(data_rows, start=2):
            if not any(_bulk_import_cell(cell) for cell in row):
                continue

            row_data = _row_to_onboarding_data(row, headers, onboarding_type)
            required_errors = _validate_required_manual_onboarding_fields(row_data)
            if required_errors:
                row_errors.append({'row': index, 'errors': required_errors})
                continue

            serializer = OnboardingDetailSerializer(
                data=row_data,
                partial=True,
                context={'require_complete': True},
            )
            if not serializer.is_valid():
                row_errors.append({'row': index, 'errors': serializer.errors})
                continue

            onboarding = serializer.save(
                created_by=request.user,
                assigned_user=None,
                status='DRAFT',
            )
            OnboardingApprovalHistory.objects.create(
                onboarding=onboarding,
                action='SUBMITTED',
                actor=request.user,
                comments='Created via bulk Excel import.',
            )
            created.append(onboarding)

        return Response(
            {
                'created_count': len(created),
                'error_count': len(row_errors),
                'created': OnboardingListSerializer(created, many=True).data,
                'errors': row_errors,
            },
            status=status.HTTP_201_CREATED if created else status.HTTP_400_BAD_REQUEST,
        )


# ── Admin: Create onboarding & send email ────────────────────────
class CreateOnboardingView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        serializer = CreateOnboardingSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        email = serializer.validated_data['email']
        onboarding_type = serializer.validated_data['onboarding_type']

        # Get or create the user for this email
        user, _ = User.objects.get_or_create(
            email=email,
            defaults={'role': onboarding_type, 'is_active': True}
        )

        onboarding = Onboarding.objects.create(
            onboarding_type=onboarding_type,
            created_by=request.user,
            assigned_user=request.user,
            emails=[email],
        )

        token = OnboardingToken.create_for_onboarding(onboarding)

        try:
            send_onboarding_invite(
                to_email=email,
                onboarding=onboarding,
                token=token.token,
            )
        except Exception as e:
            # Don't fail if email fails — log and continue
            print(f"Email send failed: {e}")

        return Response(
            {
                'onboarding': OnboardingDetailSerializer(onboarding).data,
                'token': token.token,
                'message': f'Onboarding created and invite sent to {email}',
            },
            status=status.HTTP_201_CREATED,
        )


# ── Admin: Get / update single onboarding ────────────────────────
class OnboardingDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_onboarding(self, pk):
        try:
            return Onboarding.objects.get(pk=pk)
        except Onboarding.DoesNotExist:
            return None

    def get(self, request, pk):
        onboarding = self._get_onboarding(pk)
        if not onboarding:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not _can_access(request.user, onboarding):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)
        return Response(OnboardingDetailSerializer(onboarding).data)

    def patch(self, request, pk):
        onboarding = self._get_onboarding(pk)
        if not onboarding:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        if not _can_access(request.user, onboarding):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)
        if not _can_edit_onboarding(request.user, onboarding):
            return Response({'detail': 'Only admins can edit submitted records. Employees can edit only their own drafts.'}, status=status.HTTP_403_FORBIDDEN)
        serializer = OnboardingDetailSerializer(
            onboarding,
            data=request.data,
            partial=True,
            context={'require_complete': True},
        )
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)


# ── Admin: Approve ───────────────────────────────────────────────
class ApproveOnboardingView(APIView):
    permission_classes = [IsAdminOrBoss]

    def post(self, request, pk):
        try:
            onboarding = Onboarding.objects.get(pk=pk)
        except Onboarding.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not _can_access(request.user, onboarding):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        if onboarding.status == 'APPROVED':
            return Response({'detail': 'Already approved.'}, status=status.HTTP_400_BAD_REQUEST)

        approval_messages = _validate_approval_verifications(onboarding)
        if approval_messages:
            return Response(
                {
                    'detail': approval_messages[0],
                    'messages': approval_messages,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        detail_serializer = OnboardingDetailSerializer(
            onboarding,
            data={},
            partial=True,
            context={'require_complete': True},
        )
        if not detail_serializer.is_valid():
            return Response(detail_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        document_errors = _validate_required_onboarding_documents(onboarding)
        if document_errors:
            return Response(document_errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = ApproveRejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        onboarding.status = 'APPROVED'
        onboarding.approved_by = request.user
        onboarding.approved_at = timezone.now()
        onboarding.remarks = serializer.validated_data.get('remarks', '')
        onboarding.save()
        OnboardingApprovalHistory.objects.create(
            onboarding=onboarding,
            action='APPROVED',
            actor=request.user,
            comments=onboarding.remarks,
        )
        return Response(OnboardingDetailSerializer(onboarding).data)


# ── Admin/Boss: Bulk approve ──────────────────────────────────────
class BulkApproveOnboardingView(APIView):
    permission_classes = [IsAdminOrBoss]

    def post(self, request):
        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'ids': ['Select at least one record to approve.']}, status=status.HTTP_400_BAD_REQUEST)

        remarks = request.data.get('remarks', '')

        onboardings = Onboarding.objects.filter(pk__in=ids)
        found_by_id = {str(item.id): item for item in onboardings}

        approved = []
        failed = []

        for onboarding_id in ids:
            onboarding = found_by_id.get(str(onboarding_id))
            if not onboarding:
                failed.append({'id': onboarding_id, 'errors': {'detail': 'Not found.'}})
                continue
            if not _can_access(request.user, onboarding):
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': 'Forbidden.'}})
                continue
            if onboarding.status == 'APPROVED':
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': 'Already approved.'}})
                continue

            approval_messages = _validate_approval_verifications(onboarding)
            if approval_messages:
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': approval_messages[0], 'messages': approval_messages}})
                continue

            detail_serializer = OnboardingDetailSerializer(
                onboarding,
                data={},
                partial=True,
                context={'require_complete': True},
            )
            if not detail_serializer.is_valid():
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': detail_serializer.errors})
                continue
            document_errors = _validate_required_onboarding_documents(onboarding)
            if document_errors:
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': document_errors})
                continue

            onboarding.status = 'APPROVED'
            onboarding.approved_by = request.user
            onboarding.approved_at = timezone.now()
            onboarding.remarks = remarks
            onboarding.save()
            OnboardingApprovalHistory.objects.create(
                onboarding=onboarding,
                action='APPROVED',
                actor=request.user,
                comments=remarks,
            )
            approved.append(str(onboarding.id))

        return Response({
            'approved_count': len(approved),
            'failed_count': len(failed),
            'approved': approved,
            'failed': failed,
        })


# ── Admin: Reject ────────────────────────────────────────────────
class RejectOnboardingView(APIView):
    permission_classes = [IsAdminOrBoss]

    def post(self, request, pk):
        try:
            onboarding = Onboarding.objects.get(pk=pk)
        except Onboarding.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not _can_access(request.user, onboarding):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = ApproveRejectSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        onboarding.status = 'REJECTED'
        onboarding.remarks = serializer.validated_data.get('remarks', '')
        onboarding.save()
        OnboardingApprovalHistory.objects.create(
            onboarding=onboarding,
            action='REJECTED',
            actor=request.user,
            comments=onboarding.remarks,
        )
        return Response(OnboardingDetailSerializer(onboarding).data)


# ── Admin: Re-send invite ────────────────────────────────────────
class SendToBossView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            onboarding = Onboarding.objects.get(pk=pk)
        except Onboarding.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if request.user.role != 'EMPLOYEE' or str(onboarding.created_by_id) != str(request.user.id):
            return Response({'detail': 'Only the employee who created this request can send it to a boss.'}, status=status.HTTP_403_FORBIDDEN)
        if onboarding.status == 'APPROVED':
            return Response({'detail': 'Approved onboarding cannot be sent for approval again.'}, status=status.HTTP_400_BAD_REQUEST)
        if onboarding.status == 'PENDING_BOSS_APPROVAL':
            return Response({'detail': 'This request has already been sent to a boss.'}, status=status.HTTP_400_BAD_REQUEST)

        assigned_boss, boss_error = _selected_approval_boss(request.user, request.data)
        if boss_error:
            return Response(boss_error, status=status.HTTP_400_BAD_REQUEST)

        detail_serializer = OnboardingDetailSerializer(
            onboarding,
            data={},
            partial=True,
            context={'require_complete': True},
        )
        if not detail_serializer.is_valid():
            return Response(detail_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        document_errors = _validate_required_onboarding_documents(onboarding)
        if document_errors:
            return Response(document_errors, status=status.HTTP_400_BAD_REQUEST)
        approval_messages = _validate_approval_verifications(onboarding)
        if approval_messages:
            return Response(
                {'detail': approval_messages[0], 'messages': approval_messages},
                status=status.HTTP_400_BAD_REQUEST,
            )

        onboarding.assigned_user = assigned_boss
        onboarding.status = 'PENDING_BOSS_APPROVAL'
        onboarding.save(update_fields=['assigned_user', 'status', 'updated_at'])
        OnboardingApprovalHistory.objects.create(
            onboarding=onboarding,
            action='SUBMITTED',
            actor=request.user,
            comments=f'Sent to {assigned_boss.email} for boss approval.',
        )
        _notify_boss_of_approval_request(onboarding, assigned_boss, request.user)
        return Response(OnboardingDetailSerializer(onboarding).data)


class BulkSendToBossView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request):
        if request.user.role != 'EMPLOYEE':
            return Response({'detail': 'Only employees send requests to a boss.'}, status=status.HTTP_403_FORBIDDEN)

        ids = request.data.get('ids') or []
        if not isinstance(ids, list) or not ids:
            return Response({'ids': ['Select at least one record to send.']}, status=status.HTTP_400_BAD_REQUEST)

        assigned_boss, boss_error = _selected_approval_boss(request.user, request.data)
        if boss_error:
            return Response(boss_error, status=status.HTTP_400_BAD_REQUEST)

        onboardings = Onboarding.objects.filter(pk__in=ids, created_by=request.user)
        found_by_id = {str(item.id): item for item in onboardings}

        sent = []
        failed = []

        for onboarding_id in ids:
            onboarding = found_by_id.get(str(onboarding_id))
            if not onboarding:
                failed.append({'id': onboarding_id, 'errors': {'detail': 'Not found.'}})
                continue
            if onboarding.status == 'APPROVED':
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': 'Already approved.'}})
                continue
            if onboarding.status == 'PENDING_BOSS_APPROVAL':
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': 'Already sent to a boss.'}})
                continue

            detail_serializer = OnboardingDetailSerializer(
                onboarding,
                data={},
                partial=True,
                context={'require_complete': True},
            )
            if not detail_serializer.is_valid():
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': detail_serializer.errors})
                continue
            document_errors = _validate_required_onboarding_documents(onboarding)
            if document_errors:
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': document_errors})
                continue
            approval_messages = _validate_approval_verifications(onboarding)
            if approval_messages:
                failed.append({'id': onboarding_id, 'company_name': onboarding.company_name, 'errors': {'detail': approval_messages[0], 'messages': approval_messages}})
                continue

            onboarding.assigned_user = assigned_boss
            onboarding.status = 'PENDING_BOSS_APPROVAL'
            onboarding.save(update_fields=['assigned_user', 'status', 'updated_at'])
            OnboardingApprovalHistory.objects.create(
                onboarding=onboarding,
                action='SUBMITTED',
                actor=request.user,
                comments=f'Sent to {assigned_boss.email} for boss approval.',
            )
            _notify_boss_of_approval_request(onboarding, assigned_boss, request.user)
            sent.append(str(onboarding.id))

        return Response({
            'sent_count': len(sent),
            'failed_count': len(failed),
            'sent': sent,
            'failed': failed,
        })


class ResendInviteView(APIView):
    permission_classes = [IsAdmin]

    def post(self, request, pk):
        try:
            onboarding = Onboarding.objects.get(pk=pk)
        except Onboarding.DoesNotExist:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)

        if not _can_access(request.user, onboarding):
            return Response({'detail': 'Forbidden.'}, status=status.HTTP_403_FORBIDDEN)

        token = OnboardingToken.create_for_onboarding(onboarding)
        email = onboarding.emails[0] if onboarding.emails else None
        if not email:
            return Response({'detail': 'No email on record.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            send_onboarding_invite(to_email=email, onboarding=onboarding, token=token.token)
        except Exception as e:
            return Response({'detail': f'Email failed: {e}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'detail': f'Invite resent to {email}.'})


# ── Public: Validate token & get form ────────────────────────────
class ValidateTokenView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, token):
        try:
            tok = OnboardingToken.objects.select_related('onboarding').get(token=token)
        except OnboardingToken.DoesNotExist:
            return Response({'detail': 'Invalid token.'}, status=status.HTTP_404_NOT_FOUND)

        if not tok.is_valid():
            return Response({'detail': 'Token has expired or already used.'}, status=status.HTTP_400_BAD_REQUEST)

        data = OnboardingTokenSerializer(tok).data
        data['onboarding_type'] = tok.onboarding.onboarding_type
        data['entity_type'] = _entity_type_for_onboarding(tok.onboarding)
        return Response(data)


# ── Public: Submit / update onboarding form ───────────────────────
class SubmitOnboardingView(APIView):
    permission_classes = [AllowAny]

    def put(self, request, token):
        try:
            tok = OnboardingToken.objects.select_related('onboarding').get(token=token)
        except OnboardingToken.DoesNotExist:
            return Response({'detail': 'Invalid token.'}, status=status.HTTP_404_NOT_FOUND)

        if not tok.is_valid():
            return Response({'detail': 'Token expired or already used.'}, status=status.HTTP_400_BAD_REQUEST)

        onboarding = tok.onboarding
        if onboarding.status == 'APPROVED':
            return Response({'detail': 'Approved onboarding cannot be edited.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = OnboardingDetailSerializer(onboarding, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data)

    def post(self, request, token):
        """Final submit — moves status to PENDING."""
        try:
            tok = OnboardingToken.objects.select_related('onboarding').get(token=token)
        except OnboardingToken.DoesNotExist:
            return Response({'detail': 'Invalid token.'}, status=status.HTTP_404_NOT_FOUND)

        if not tok.is_valid():
            return Response({'detail': 'Token expired or already used.'}, status=status.HTTP_400_BAD_REQUEST)

        onboarding = tok.onboarding
        if onboarding.status == 'APPROVED':
            return Response({'detail': 'Approved onboarding cannot be re-submitted.'}, status=status.HTTP_403_FORBIDDEN)

        serializer = OnboardingDetailSerializer(
            onboarding,
            data=request.data,
            partial=True,
            context={'require_complete': True},
        )
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        instance = serializer.save(status='PENDING', assigned_user=onboarding.created_by)
        OnboardingApprovalHistory.objects.create(
            onboarding=instance,
            action='SUBMITTED',
            actor=None,
            comments='Submitted by vendor/customer to the link sender for review.',
        )
        document_errors = _validate_required_onboarding_documents(instance)
        if document_errors:
            return Response(document_errors, status=status.HTTP_400_BAD_REQUEST)
        return Response(OnboardingDetailSerializer(instance).data)


# ── Admin: Dashboard stats ────────────────────────────────────────
class DashboardStatsView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        base = _scoped_qs(request.user)
        try:
            base = _filter_by_created_date(base, request.query_params)
        except ValueError as error:
            return Response(error.args[0], status=status.HTTP_400_BAD_REQUEST)

        qs = base
        onboarding_type = request.query_params.get('type')
        if onboarding_type:
            qs = qs.filter(onboarding_type=onboarding_type.upper())

        stats = {
            'total':    qs.count(),
            'pending':  qs.filter(status__in=PENDING_STATUS_FILTER).count(),
            'approved': qs.filter(status='APPROVED').count(),
            'rejected': qs.filter(status='REJECTED').count(),
            'msme':     qs.filter(msme_applicable=True).count(),
            'vendor':   base.filter(onboarding_type='VENDOR').count(),
            'customer': base.filter(onboarding_type='CUSTOMER').count(),
            'employees': request.user.employees.count() if request.user.role == 'BOSS' else 0,
        }
        return Response(stats)


class VendorReferenceRangeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        ranges = VendorReferenceMaster.objects.order_by('vendor_reference_range').values_list(
            'vendor_reference_range',
            flat=True,
        )
        return Response([
            {'value': value, 'label': value.replace('-', ' - ')}
            for value in ranges
        ])


class PaymentTermListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        terms = PaymentTermMaster.objects.order_by('payment_term')
        return Response([
            {
                'value': term.payment_term,
                'label': f'{term.payment_term} - {term.description}',
                'description': term.description,
            }
            for term in terms
        ])


class PurchaseOrganizationListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        organizations = PurchaseOrganizationMaster.objects.order_by('purchase_organization')
        return Response([
            {
                'value': organization.purchase_organization,
                'label': f'{organization.purchase_organization} - {organization.description}',
                'description': organization.description,
            }
            for organization in organizations
        ])


class CompanyCodeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_codes = CompanyCodeMaster.objects.order_by('company_code')
        return Response([
            {
                'value': company_code.company_code,
                'label': f'{company_code.company_code} - {company_code.name}',
                'name': company_code.name,
            }
            for company_code in company_codes
        ])


class TDSCodeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tds_codes = TDSCodeMaster.objects.order_by('tds_code')
        return Response([
            {
                'value': tds_code.tds_code,
                'label': f'{tds_code.tds_code} - {tds_code.description}',
                'description': tds_code.description,
            }
            for tds_code in tds_codes
        ])


class SearchTermListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        search_terms = SearchTermMaster.objects.order_by('search_term')
        return Response([
            {
                'value': search_term.search_term,
                'label': f'{search_term.search_term} - {search_term.applicable_for}',
                'applicable_for': search_term.applicable_for,
            }
            for search_term in search_terms
        ])


class SalesOrganizationListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sales_organizations = SalesOrganizationMaster.objects.order_by('sales_organization')
        return Response([
            {
                'value': sales_organization.sales_organization,
                'label': f'{sales_organization.sales_organization} - {sales_organization.description}',
                'description': sales_organization.description,
            }
            for sales_organization in sales_organizations
        ])


class DistributionChannelListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        distribution_channels = DistributionChannelMaster.objects.order_by('distribution_channel')
        return Response([
            {
                'value': distribution_channel.distribution_channel,
                'label': f'{distribution_channel.distribution_channel} - {distribution_channel.description}',
                'description': distribution_channel.description,
            }
            for distribution_channel in distribution_channels
        ])


class DivisionListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        divisions = DivisionMaster.objects.order_by('division')
        return Response([
            {
                'value': division.division,
                'label': f'{division.division} - {division.description}',
                'description': division.description,
            }
            for division in divisions
        ])


class TransportationZoneListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        zones = TransportationZoneMaster.objects.order_by('zone')
        return Response([
            {
                'value': zone.zone,
                'label': f'{zone.zone} - {zone.description}',
                'description': zone.description,
            }
            for zone in zones
        ])


class CustomerCompanyCodeListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        company_codes = CustomerCompanyCodeMaster.objects.order_by('company_code')
        return Response([
            {
                'value': company_code.company_code,
                'label': f'{company_code.company_code} - {company_code.name}',
                'name': company_code.name,
            }
            for company_code in company_codes
        ])


class VendorReferenceMasterListCreateView(APIView):
    permission_classes = [IsFullAdmin]

    def get(self, request):
        qs = VendorReferenceMaster.objects.all()
        search = request.query_params.get('search', '').strip()
        if search:
            qs = qs.filter(
                Q(vendor_reference_range__icontains=search) |
                Q(group_code__icontains=search) |
                Q(nr_group__icontains=search) |
                Q(reference_name__icontains=search) |
                Q(gl_account_number__icontains=search) |
                Q(gl_account_description__icontains=search)
            )
        return Response(VendorReferenceMasterSerializer(qs, many=True).data)

    def post(self, request):
        serializer = VendorReferenceMasterSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        master = serializer.save()
        return Response(VendorReferenceMasterSerializer(master).data, status=status.HTTP_201_CREATED)


class VendorReferenceMasterDetailView(APIView):
    permission_classes = [IsFullAdmin]

    def _get_master(self, pk):
        try:
            return VendorReferenceMaster.objects.get(pk=pk)
        except VendorReferenceMaster.DoesNotExist:
            return None

    def get(self, request, pk):
        master = self._get_master(pk)
        if not master:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(VendorReferenceMasterSerializer(master).data)

    def put(self, request, pk):
        master = self._get_master(pk)
        if not master:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = VendorReferenceMasterSerializer(master, data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        master = serializer.save()
        return Response(VendorReferenceMasterSerializer(master).data)

    def patch(self, request, pk):
        master = self._get_master(pk)
        if not master:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = VendorReferenceMasterSerializer(master, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        master = serializer.save()
        return Response(VendorReferenceMasterSerializer(master).data)

    def delete(self, request, pk):
        master = self._get_master(pk)
        if not master:
            return Response({'detail': 'Not found.'}, status=status.HTTP_404_NOT_FOUND)
        master.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class VendorReferenceLookupView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = VendorReferenceLookupSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        vendor_reference_range = serializer.validated_data.get('vendor_reference_range')
        if not vendor_reference_range:
            vendor_reference_range = get_vendor_reference_range_for_code(
                serializer.validated_data.get('vendor_reference_code')
            )
        if not vendor_reference_range:
            return Response(
                {'detail': 'Vendor reference code does not fall in a predefined range.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        try:
            master = VendorReferenceMaster.objects.get(vendor_reference_range=vendor_reference_range)
        except VendorReferenceMaster.DoesNotExist:
            return Response(
                {
                    'detail': 'No Vendor Reference Master mapping exists for the matched range.',
                    'vendor_reference_range': vendor_reference_range,
                },
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(VendorReferenceMasterSerializer(master).data)











from rest_framework.views import APIView
from rest_framework.response import Response

# class VerifyPANAPIView(APIView):

#     def post(self, request):

#         pan = request.data.get("pan_number")
#         dob = request.data.get("date_of_birth")

#         return Response({
#             "status": True,
#             "message": "PAN verification API called successfully",
#             "pan": pan,
#             "dob": dob
#         })
    

import requests

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .utils.sandbox import get_access_token
from datetime import datetime
import requests
# views.py

from datetime import datetime

import requests

from django.conf import settings
from django.utils import timezone

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import Onboarding
from .utils.sandbox import get_access_token


class VerifyPANAPIView(APIView):

    def post(self, request):

        onboarding_id = request.data.get("onboarding_id")

        pan = request.data.get("pan_number")
        dob = request.data.get("date_of_birth")
        name = request.data.get("name")

        if not pan:
            return Response(
                {"error": "PAN Number is required"},
                status=400
            )

        try:

            onboarding = Onboarding.objects.get(
                id=onboarding_id
            )

            dob = datetime.strptime(
                dob,
                "%Y-%m-%d"
            ).strftime("%d/%m/%Y")

            token = get_access_token()

            payload = {
                "@entity": "in.co.sandbox.kyc.pan_verification.request",
                "pan": pan.upper(),
                "name_as_per_pan": name,
                "date_of_birth": dob,
                "consent": "Y",
                "reason": "Vendor onboarding PAN verification"
            }

            response = requests.post(
                "https://api.sandbox.co.in/kyc/pan/verify",
                json=payload,
                headers={
                    "Authorization": token,
                    "x-api-key": settings.SANDBOX_API_KEY,
                    "x-api-version": "1.0.0",
                    "Content-Type": "application/json"
                },
                timeout=30
            )

            result = response.json()

            pan_data = result.get("data", {})

            is_verified = (
                pan_data.get("status") == "valid"
            )

            if is_verified:

                if pan_data.get(
                    "aadhaar_seeding_status"
                ) == "y":

                    verification_status = (
                        "Valid and Operative (Aadhaar Linked)"
                    )

                else:

                    verification_status = (
                        "Valid and Operative"
                    )

            else:

                verification_status = (
                    pan_data.get(
                        "remarks",
                        "Invalid PAN"
                    )
                )

            onboarding.pan_verified = is_verified

            onboarding.pan_verification_status = (
                verification_status
            )

            onboarding.pan_verification_response = (
                result
            )

            onboarding.save()

            return Response({
                "verified": is_verified,
                "verification_status":
                    verification_status,
                "sandbox_response":
                    result
            })

        except Onboarding.DoesNotExist:

            return Response(
                {
                    "error":
                    "Onboarding record not found"
                },
                status=404
            )

        except Exception as e:

            return Response(
                {
                    "error": str(e)
                },
                status=400
            )


class VerifyGSTAPIView(APIView):

    def post(self, request):

        onboarding_id = request.data.get("onboarding_id")
        gstin = request.data.get("gst_number")

        if not gstin:
            return Response(
                {"error": "GST Number is required"},
                status=400
            )

        try:
            onboarding = Onboarding.objects.get(id=onboarding_id)
            expected_state_code = gst_state_code_for_state(onboarding.state)
            gst_state_code = gstin.upper()[:2]
            if expected_state_code and gst_state_code != expected_state_code:
                return Response(
                    {
                        "error": (
                            f"GST first two digits must be {expected_state_code} "
                            f"for {onboarding.state}"
                        )
                    },
                    status=400
                )

            token = get_access_token()

            payload = {"gstin": gstin.upper()}

            response = requests.post(
                "https://api.sandbox.co.in/gst/compliance/public/gstin/search",
                json=payload,
                headers={
                    "Authorization": token,
                    "x-api-key": settings.SANDBOX_API_KEY,
                    "x-api-version": "1.0.0",
                    "Content-Type": "application/json"
                },
                timeout=30
            )

            result = response.json()

            # Sandbox nests the GSTIN payload under data -> data
            outer = result.get("data", {}) or {}
            gst_data = outer.get("data", {}) or {}

            found = outer.get("status_cd") == "1" and bool(gst_data)
            sts = gst_data.get("sts")  # "Active", "Cancelled", "Suspended"

            is_verified = found and sts == "Active"

            if is_verified:
                verification_status = (
                    f"Valid and Active — {gst_data.get('lgnm', '')}".strip(" —")
                )
            elif found:
                verification_status = f"GSTIN found but status is {sts}"
            else:
                verification_status = "Invalid GSTIN / No records found"

            onboarding.gst_verified = is_verified
            onboarding.gst_verification_status = verification_status
            onboarding.gst_verification_response = result
            onboarding.save()

            return Response({
                "verified": is_verified,
                "verification_status": verification_status,
                "sandbox_response": result
            })

        except Onboarding.DoesNotExist:
            return Response(
                {"error": "Onboarding record not found"},
                status=404
            )
