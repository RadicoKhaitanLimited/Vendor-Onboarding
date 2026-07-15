from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.onboarding.models import DeliveryPlantMaster


class Command(BaseCommand):
    help = (
        'Load Delivery Plant master data from the "Customer Shipping - Delivery Plants" '
        'excel export (columns: SOrg., DChl, Plant, Plant Name). '
        'Existing rows are updated in place; new plants are created.'
    )

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to the Delivery Plants .xlsx file')
        parser.add_argument(
            '--sheet', type=str, default=None,
            help='Worksheet name to read (defaults to the first sheet)',
        )

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        try:
            workbook = load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f'File not found: {excel_path}')

        sheet_name = options['sheet']
        worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

        created, updated, skipped = 0, 0, 0
        for row in worksheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 4:
                continue
            sales_org, dist_channel, plant, plant_name = row[0], row[1], row[2], row[3]
            if not plant or str(plant).strip().upper() in ('PLANT', ''):
                continue

            plant_code = str(plant).strip().upper()
            defaults = {
                'plant_name': str(plant_name or '').strip(),
                'sales_organization': str(sales_org or '').strip().upper(),
                'distribution_channel': str(dist_channel or '').strip().upper(),
            }
            obj, was_created = DeliveryPlantMaster.objects.update_or_create(
                plant=plant_code, defaults=defaults,
            )
            if was_created:
                created += 1
            else:
                updated += 1

        self.stdout.write(self.style.SUCCESS(
            f'Delivery plants loaded: {created} created, {updated} updated, {skipped} skipped.'
        ))
